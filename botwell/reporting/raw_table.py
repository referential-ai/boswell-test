"""
Raw numeric table generation for Boswell Test results.

This module generates a simplified Excel spreadsheet with raw numeric grades
without formatting, colors, or descriptive text.
"""

import os
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from botwell.reporting.excel import abbreviate_model_name

# Define minimal styling constants
THIN_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'),
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
BOLD_FONT = Font(name="Helvetica Neue", bold=True, size=8)
REGULAR_FONT = Font(name="Helvetica Neue", size=8)
CENTER_ALIGNMENT = Alignment(horizontal='center')

def generate_raw_numeric_table(results: Dict[str, Any], output_path: str) -> None:
    """
    Generate a simplified Excel table with only raw numeric grades.
    
    Args:
        results: The full results dictionary with grades
        output_path: Path where to save the Excel file
    """
    try:
        # Extract models list
        models = list(results["essays"].keys())
        if not models:
            raise ValueError("No models found in results")
            
        # Abbreviate model names for display
        display_models = [abbreviate_model_name(model) for model in models]
        
        # Create workbook and get active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Raw Numeric Data"
        
        # Set up header row
        ws.cell(row=1, column=1, value="Model").font = BOLD_FONT
        ws.cell(row=1, column=1).border = THIN_BORDER
        
        # Add model columns
        for j, model in enumerate(display_models):
            cell = ws.cell(row=1, column=j+2, value=model)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            # Set narrow column width for compact display
            ws.column_dimensions[get_column_letter(j+2)].width = 10
            
        # Add median and average columns
        median_col = len(models) + 2
        avg_col = len(models) + 3
        ws.cell(row=1, column=median_col, value="Median").font = BOLD_FONT
        ws.cell(row=1, column=median_col).border = THIN_BORDER
        ws.cell(row=1, column=avg_col, value="Average").font = BOLD_FONT
        ws.cell(row=1, column=avg_col).border = THIN_BORDER
        
        # Fill in raw data rows
        for i, graded_model in enumerate(models):
            row = i + 2  # Start on row 2
            
            # Add model name in first column
            model_cell = ws.cell(row=row, column=1, value=display_models[i])
            model_cell.font = BOLD_FONT
            model_cell.border = THIN_BORDER
            
            # Collect all raw numeric grades for this model
            raw_numeric_grades = []
            
            # Add each grade as raw numeric value
            for j, grader_model in enumerate(models):
                col = j + 2
                
                # Extract grade data
                try:
                    grade_info = results["grades"][graded_model][grader_model]
                    if isinstance(grade_info, dict) and "numeric_grade" in grade_info:
                        raw_grade = grade_info["numeric_grade"]
                        if raw_grade > 0:
                            # Add raw numeric grade
                            ws.cell(row=row, column=col, value=round(raw_grade, 2))
                            raw_numeric_grades.append(raw_grade)
                except (KeyError, TypeError):
                    # Leave cell empty if no grade data
                    pass
                
                # Add border to all cells
                ws.cell(row=row, column=col).border = THIN_BORDER
            
            # Calculate and add median (if enough data)
            if raw_numeric_grades:
                sorted_grades = sorted(raw_numeric_grades)
                # Calculate median based on number of grades
                if len(sorted_grades) % 2 == 0:  # Even number
                    median = (sorted_grades[len(sorted_grades)//2 - 1] + sorted_grades[len(sorted_grades)//2]) / 2
                else:  # Odd number
                    median = sorted_grades[len(sorted_grades)//2]
                    
                # Add median to table
                median_cell = ws.cell(row=row, column=median_col, value=round(median, 2))
                median_cell.border = THIN_BORDER
                
                # Calculate and add average
                avg = sum(sorted_grades) / len(sorted_grades)
                avg_cell = ws.cell(row=row, column=avg_col, value=round(avg, 2))
                avg_cell.border = THIN_BORDER
            else:
                # Add empty cells with borders for median and average
                ws.cell(row=row, column=median_col).border = THIN_BORDER
                ws.cell(row=row, column=avg_col).border = THIN_BORDER
        
        # Add column median row at the bottom
        col_median_row = len(models) + 2
        median_label = ws.cell(row=col_median_row, column=1, value="Median Grade Given")
        median_label.font = BOLD_FONT
        median_label.border = THIN_BORDER
        
        # Calculate and add column medians
        for j, grader_model in enumerate(models):
            col = j + 2
            
            # Collect all numeric grades given by this model
            col_grades = []
            for graded_model in models:
                try:
                    if graded_model in results["grades"] and grader_model in results["grades"][graded_model]:
                        grade_info = results["grades"][graded_model][grader_model]
                        if isinstance(grade_info, dict) and "numeric_grade" in grade_info:
                            numeric_grade = grade_info["numeric_grade"]
                            if numeric_grade > 0:
                                col_grades.append(numeric_grade)
                except KeyError:
                    pass
            
            # Calculate and add median for this column
            if col_grades:
                col_grades.sort()
                if len(col_grades) % 2 == 0:  # Even number
                    col_median = (col_grades[len(col_grades)//2 - 1] + col_grades[len(col_grades)//2]) / 2
                else:  # Odd number
                    col_median = col_grades[len(col_grades)//2]
                
                # Add to the cell
                ws.cell(row=col_median_row, column=col, value=round(col_median, 2))
            
            # Apply border to all cells in the row
            ws.cell(row=col_median_row, column=col).border = THIN_BORDER
        
        # Add borders to median and average cells in bottom row
        ws.cell(row=col_median_row, column=median_col).border = THIN_BORDER
        ws.cell(row=col_median_row, column=avg_col).border = THIN_BORDER
        
        # Save the Excel file
        excel_file = os.path.join(output_path, "raw_numeric_table.xlsx")
        wb.save(excel_file)
        # Return the path but don't print - let the calling script handle output
        
        return excel_file
        
    except Exception as e:
        print(f"Error generating raw numeric table: {str(e)}")
        if isinstance(e, (KeyError, IndexError, ValueError, TypeError)):
            print(f"Details: {type(e).__name__} - Check the data structure")
        return None