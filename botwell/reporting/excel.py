"""
Excel cross-assessment table generation for Boswell Test results.

This module generates a color-coded Excel spreadsheet with the complete
cross-assessment matrix as seen in Table 1 of the research paper.
"""

import os
import sys
from typing import Dict, Any, List, Optional, Union, Tuple
import re
import json
from dataclasses import dataclass
import openpyxl 
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.fonts import DEFAULT_FONT
from openpyxl.worksheet.worksheet import Worksheet
from botwell.core.grading import numeric_to_composite_grade
from openpyxl.cell import Cell

# Letter grades to numeric scores mapping (as shown in the image)
GRADE_SCORES = {
    "A+": 97.0,
    "A": 93.0,
    "A-": 90.0,
    "B+": 87.0,
    "B": 83.0,
    "B-": 80.0,
    "C+": 77.0,
    "C": 73.0,
    "C-": 70.0,
    "D+": 67.0,
    "D": 63.0,
    "D-": 60.0,
    "F": 50.0,
    "N/A": 0
}

# GPA-style grades (0-4 scale) mapping, for detecting GPA-scale scores
GPA_GRADE_SCALE = {
    "A+": 4.0,
    "A": 4.0,
    "A-": 3.7,
    "B+": 3.3,
    "B": 3.0,
    "B-": 2.7,
    "C+": 2.3,
    "C": 2.0,
    "C-": 1.7,
    "D+": 1.3,
    "D": 1.0,
    "D-": 0.7,
    "F": 0.0,
    "N/A": 0
}

# Define colors for different grades (exactly as in the image)
GRADE_COLORS = {
    "A+": "00FF00",  # Bright green
    "A": "90EE90",   # Light green
    "A-": "98FB98",  # Pale green
    "B+": "ADD8E6",  # Light blue
    "B": "87CEEB",   # Sky blue
    "B-": "B0C4DE",  # Light steel blue
    "C+": "FFFFE0",  # Light yellow
    "C": "FFFF99",   # Yellow
    "C-": "FFD700",  # Gold
    "D+": "FFA07A",  # Light salmon
    "D": "FF7F50",   # Coral
    "D-": "FF6347",  # Tomato
    "F": "FF0000",   # Red
    "N/A": "D3D3D3", # Light gray
}

# Define common styles as module-level constants
THIN_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'),
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
CENTER_ALIGNMENT = Alignment(horizontal='center')
RIGHT_ALIGNMENT = Alignment(horizontal='right')
LEFT_ALIGNMENT = Alignment(horizontal='left')
MODEL_NAME_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray
HEADER_ROW_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Same gray for headers
BIAS_ROW_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
HELVETICA_FONT = Font(name="Helvetica Neue", size=8)  # All text at 8pt
HELVETICA_BOLD = Font(name="Helvetica Neue", bold=True, size=8)  # Bold 8pt for header/footer rows
# Larger font for model names (3 points larger as requested)
MODEL_NAME_FONT = Font(name="Helvetica Neue", bold=True, size=11)  # 11pt font for model names
WRAP_ALIGNMENT = Alignment(horizontal='left', wrap_text=True)  
CENTER_WRAP_ALIGNMENT = Alignment(horizontal='center', wrap_text=True)  # Use for all cells

# Fill color for self-assessment cells
SELF_ASSESSMENT_FILL = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light lavender

def abbreviate_model_name(model_name: str) -> str:
    """
    Create shortened version of model name for display.
    
    Args:
        model_name: Full model name
        
    Returns:
        Abbreviated model name
    """
    # Common patterns to abbreviate
    patterns = [
        # DeepSeek models
        (r'DeepSeek-Distill-Qwen-(\d+b)', r'DeepSeek \1'),
        (r'DeepSeek-R1-Full', r'DeepSeek R1'),
        # Perplexity models
        (r'Perplexity: Llama 3.1 Sonar (\d+B)(?: Online)?', r'Perplexity \1'),
        # Claude models
        (r'Claude-(\d+(?:\.\d+)?)-(\w+)', r'Claude \1 \2'),
        # Gemini models
        (r'Gemini (Pro|Flash) (\d+\.\d+)', r'Gemini \1 \2'),
        # GPT models
        (r'GPT-(\d+)([a-z]?)(?:-mini)?', r'GPT-\1\2'),
        # Others 
        (r'grok(\d+)-(\d+)', r'Grok \1'),
        (r'o(\d+)(?:-mini)?(?:-high)?', r'o\1')
    ]
    
    # Apply patterns
    result = model_name
    for pattern, replacement in patterns:
        result = re.sub(pattern, replacement, result)
    
    return result

def apply_cell_style(
    cell: Cell, 
    value: Optional[str] = None, 
    alignment: Optional[Alignment] = None, 
    border: Optional[Border] = None,
    font: Optional[Font] = HELVETICA_FONT,
    fill: Optional[PatternFill] = None
) -> Cell:
    """
    Apply common styling to a cell.
    
    Args:
        cell: The cell to style
        value: Optional value to set
        alignment: Optional alignment style
        border: Optional border style (defaults to THIN_BORDER)
        font: Optional font style (defaults to HELVETICA_FONT)
        fill: Optional fill style
        
    Returns:
        The styled cell
    """
    if value is not None:
        cell.value = value
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    return cell

def setup_worksheet_header(
    ws: Worksheet, 
    models: List[str]
) -> Tuple[int, int, int]:
    """
    Set up the worksheet title and header row.
    
    Args:
        ws: The worksheet to modify
        models: List of model names
        
    Returns:
        Tuple of (median_column_index, header_row_index)
    """
    # Add clearer title
    title_cell = ws.cell(row=1, column=1, value="Cross-Grading Table: Models Evaluating Each Other's Essays")
    title_cell.alignment = CENTER_ALIGNMENT
    
    # Add a note in the top right explaining the meaning of the columns (numeric values)
    note_cell = ws.cell(row=1, column=len(models)+2, value="Note: All values shown are raw numeric scores (0-4.25 scale)")
    note_cell.alignment = RIGHT_ALIGNMENT
    note_cell.font = HELVETICA_FONT
    
    # Use bold font for title
    title_cell.font = HELVETICA_BOLD
    
    # Merge cells for title
    end_col = len(models) + 3  # +2 for model column, median column, and numeric average column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Model column
    for i in range(len(models)):  # Use consistent width for all data columns
        ws.column_dimensions[get_column_letter(i+2)].width = 12
    # Add column for raw numeric average
    ws.column_dimensions[get_column_letter(len(models) + 3)].width = 12  # Raw average column
    
    # Add clearer "Model Being Evaluated" header
    apply_cell_style(
       ws.cell(row=2, column=1),
       value="Model Being Evaluated (Essay Author)",
       alignment=CENTER_WRAP_ALIGNMENT,
       border=THIN_BORDER,
       fill=MODEL_NAME_FILL,
       font=HELVETICA_BOLD
    )
    
    # Add clearer model headers with "Score from [Model Name]"
    for i, model in enumerate(models):
        # Handle repeated model names if present
        model_name = model
        # Get abbreviated name
        abbrev_name = abbreviate_model_name(model)
        
        if models.count(model) > 1:
            abbrev_name = f"{abbrev_name} (Run {models[:i+1].count(model)})"
        
        apply_cell_style(
            ws.cell(row=2, column=i+2),
            value=f"{abbrev_name}",  # Removed "Score from" as requested
            alignment=CENTER_WRAP_ALIGNMENT,
            border=THIN_BORDER,
            font=HELVETICA_BOLD,
            fill=HEADER_ROW_FILL
        )
    
    # Add Median Grade header
    median_col = len(models) + 2
    apply_cell_style(
        ws.cell(row=2, column=median_col),
        value="Median Grade",
        alignment=CENTER_WRAP_ALIGNMENT,
        border=THIN_BORDER,
        font=HELVETICA_BOLD,
        fill=HEADER_ROW_FILL
    )
    
    # Add Raw Average header
    raw_avg_col = len(models) + 3
    apply_cell_style(
        ws.cell(row=2, column=raw_avg_col),
        value="Overall Score (0-4.25 scale)",
        alignment=CENTER_WRAP_ALIGNMENT,
        border=THIN_BORDER,
        font=HELVETICA_BOLD,
        fill=PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")  # Slightly different blue
        )
        
    return median_col, raw_avg_col, 2  # Return median column index, raw average column index, and header row index

def get_grade_with_score(
    results: Dict[str, Any],
    graded_model: str,
    grader_model: str
) -> Tuple[str, str]:
    """
    Get grade and formatted grade value with score.
    
    Args:
        results: Results dictionary
        graded_model: Model being graded
        grader_model: Model doing the grading
        
    Returns:
        Tuple of (grade, formatted_grade_value, raw_numeric_grade)
    """
    try:
        grade = results["grades"][graded_model][grader_model]["grade"]
        # Use the actual numeric grade from results if available, otherwise fall back to mapping
        if "numeric_grade" in results["grades"][graded_model][grader_model]:
            score = results["grades"][graded_model][grader_model]["numeric_grade"]
            raw_numeric_grade = score
        else:
            score = GRADE_SCORES.get(grade, 0)
            raw_numeric_grade = score
        
        # Display only the raw numeric score as requested by Peter
        grade_value = f"{raw_numeric_grade:.2f}" if raw_numeric_grade > 0 else "N/A"
        
        return grade, grade_value, raw_numeric_grade
    except KeyError:
        return "N/A", "N/A", 0.0  # No numeric score

def calculate_median_grade(
    grades: List[str],
    results: Dict[str, Any] = None,
    graded_model: str = None,
    grader_models: List[str] = None
) -> Tuple[str, float, str]:
    """
    Calculate median grade from a list of grades.
    
    Args:
        grades: List of grade letters
        results: Optional results dictionary containing numeric grades
        graded_model: Optional model being graded
        grader_models: Optional list of models who gave the grades (aligned with grades list)
        
    Returns:
        Tuple of (median_grade_letter, median_score, formatted_median_value, raw_average)
    """
    # Track all raw numeric grades for computing mean
    raw_numeric_grades = []
    
    if not grades:
        return "N/A", 0, "N/A (0.00)", 0.0  # Consistent decimal formatting
    
    # Convert letter grades to numeric scores, using actual numeric grades when available
    scores = []
    # If we have all the necessary information and grader models, get exact scores
    if results and graded_model and grader_models and len(grades) == len(grader_models):
        for i, grade in enumerate(grades):
            grader_model = grader_models[i]
            # Check if this grader has graded this model
            if (grader_model in results["grades"] and 
                graded_model in results["grades"][grader_model] and
                "grade" in results["grades"][grader_model][graded_model]):
                grade_data = results["grades"][grader_model][graded_model]
                if "numeric_grade" in grade_data:
                    scores.append(grade_data["numeric_grade"])
                    raw_numeric_grades.append(grade_data["numeric_grade"])
                else:
                    scores.append(GRADE_SCORES.get(grade, 0))
                    raw_numeric_grades.append(GRADE_SCORES.get(grade, 0))
    # If no scores were found via results (or results not provided), use the mapping
    if not scores:
        scores = [GRADE_SCORES.get(grade, 0) for grade in grades]
        raw_numeric_grades = scores.copy()
    
    scores.sort()
    
    # Calculate median score (could be a decimal if even number of elements)
    length = len(scores)
    if length % 2 == 0:  # Even number, average the two middle values
        median_score = (scores[length//2 - 1] + scores[length//2]) / 2
    else:  # Odd number, take the middle value
        median_score = scores[length//2]
    
    # Detect if we're using a 0-4 scale (GPA scale) instead of 0-100 scale
    is_gpa_scale = all(score <= 4.0 and score > 0 for score in scores if score > 0)
    
    # Find the closest letter grade for the median score
    median_grade = "N/A"
    
    if is_gpa_scale:
        # For GPA scale (0-4), use direct mapping rather than ranges
        if median_score >= 4.0:
            median_grade = "A+"
        elif median_score >= 3.7:
            median_grade = "A-"
        elif median_score >= 3.3:
            median_grade = "B+"
        elif median_score >= 3.0:
            median_grade = "B"
        elif median_score >= 2.7:
            median_grade = "B-"
        elif median_score >= 2.3:
            median_grade = "C+"
        elif median_score >= 2.0:
            median_grade = "C"
        elif median_score >= 1.7:
            median_grade = "C-"
        elif median_score >= 1.3:
            median_grade = "D+"
        elif median_score >= 1.0:
            median_grade = "D"
        elif median_score >= 0.7:
            median_grade = "D-"
        elif median_score > 0:
            median_grade = "F"
    else:
        # Sort grades by score in descending order, but exclude "N/A"
        sorted_grades = sorted(
            [(grade, score) for grade, score in GRADE_SCORES.items() if grade != "N/A"],
            key=lambda x: x[1],
            reverse=True
        )
        
        for grade, score in sorted_grades:
            if median_score >= score - 0.0001:  # Add small epsilon to handle floating point comparison
                median_grade = grade
                break
    
    # Special case: if the score is very close to 0, it should be N/A
    if abs(median_score) < 0.0001:
        median_grade = "N/A"

    # Calculate raw average
    raw_average = sum(raw_numeric_grades) / len(raw_numeric_grades) if raw_numeric_grades else 0.0
    
    # For display, use the raw score instead of the letter grade
    median_value = f"{median_score:.2f}" if median_score > 0 else "N/A"
    
    return median_grade, median_score, median_value, raw_average

def extract_raw_grade_matrix(results: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Extract raw grade matrix for future reuse.
    
    Args:
        results: Full results dictionary
        
    Returns:
        Dictionary with raw grade matrix
    """
    models = list(results["essays"].keys())
    matrix = {}
    
    for graded_model in models:
        matrix[graded_model] = {}
        for grader_model in models:
            try:
                grade_info = results["grades"][graded_model][grader_model]
                matrix[graded_model][grader_model] = {
                    "letter_grade": grade_info.get("grade", "N/A"),
                    "numeric_grade": grade_info.get("numeric_grade", 0.0)
                }
            except KeyError:
                matrix[graded_model][grader_model] = {"letter_grade": "N/A", "numeric_grade": 0.0}
    
    return matrix

def handle_duplicate_model_names(models: List[str]) -> List[str]:
    """
    Process model names to handle duplicates by adding run numbers.
    
    Args: 
        models: List of model names, potentially containing duplicates
        
    Returns:
        List of model names with run numbers added to duplicates
    """
    # Create a dictionary to track occurrences of each model name 
    model_counts = {}
    processed_models = []
    
    for model in models:
        if model in model_counts:
            model_counts[model] += 1
            processed_models.append(f"{model} (Run {model_counts[model]})")
        else:
            model_counts[model] = 1
            processed_models.append(model)
    
    return processed_models

def generate_cross_grading_excel(results: Dict[str, Any], run_dir: str) -> None:
    """
    Generate Excel spreadsheet with color-coded cross-grading matrix.
    
    Args:
        results: The full results dictionary
        run_dir: Directory path where output files should be saved
    """
    try:
        # Validate inputs
        if not isinstance(results, dict) or "essays" not in results:
            raise ValueError("Results dictionary is invalid or missing essays")
        
        if not os.path.isdir(run_dir):
            os.makedirs(run_dir, exist_ok=True)
            
        # Get models list
        models = list(results["essays"].keys())
        
        # Process models list to handle duplicates
        # First abbreviate all model names
        display_models = [abbreviate_model_name(model) for model in models]

        # Then handle duplicates
        if len(set(models)) < len(models):
            display_models = handle_duplicate_model_names(models)
            
        if not display_models:
            raise ValueError("No models found in results")
        
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cross Grading"
        
        # Set up worksheet header
        median_col, raw_avg_col, header_row = setup_worksheet_header(ws, display_models)
        
        # Apply Helvetica Neue font to all cells we'll be working with
        for row in range(1, len(models) + 5):  # Add a few extra rows 
            for col in range(1, len(models) + 4):  # Add columns including raw average column
                cell = ws.cell(row=row, column=col)
                if cell.font.name != "Helvetica Neue":  # Only set if not already set
                    cell.font = HELVETICA_FONT
        
        # Fill in the grading matrix
        model_name_mapping = {}  # For appendix
        
        for i, graded_model in enumerate(models):
            row = i + 3  # Start on row 3

            # Get display name for this model (with run number if it's a duplicate)
            display_graded_model = display_models[i]
            model_name_mapping[display_graded_model] = graded_model  # Store for appendix
            
            # Add model name in first column with clearer formatting
            display_model_name = display_graded_model
            if "Run" in display_graded_model:
                model_font = Font(name="Helvetica Neue", bold=True, size=8, italic=True)
                
            apply_cell_style(
                ws.cell(row=row, column=1),
                value=display_model_name,
                alignment=WRAP_ALIGNMENT,
                border=THIN_BORDER,
                fill=MODEL_NAME_FILL,
                font=model_font if "Run" in display_graded_model else MODEL_NAME_FONT  # Use larger font for model names
            )
            
            # Add grades for each grader
            grades_for_model = []
            grader_models_for_median = []  # Track which graders provided the grades
            raw_numeric_grades = []  # Track raw numeric grades for this model
            
            for j, grader_model in enumerate(models):
                # Get display name for this grader (with run number if it's a duplicate)
                display_grader_model = display_models[j]
                col = j + 2
                
                # Get the grade that grader_model gave to graded_model (including self-assessment)
                grade, grade_value, raw_numeric_grade = get_grade_with_score(results, graded_model, grader_model)
                
                if grade != "N/A":
                    grades_for_model.append(grade)
                    grader_models_for_median.append(grader_model)
                    raw_numeric_grades.append(raw_numeric_grade)
                
                # Add the grade to the cell
                cell = ws.cell(row=row, column=col)
                # Show both letter grade and raw numeric score for more clarity
                # Display composite grades when applicable
                display_value = "N/A" 
                if grade != "N/A" and raw_numeric_grade > 0:
                    # Get composite grade if applicable
                    composite_grade = numeric_to_composite_grade(raw_numeric_grade)
                    display_value = composite_grade
                
                apply_cell_style(
                    cell,
                    value=display_value,
                    alignment=CENTER_WRAP_ALIGNMENT,
                    border=THIN_BORDER
                )
                
                # Set row height to accommodate two lines of text
                ws.row_dimensions[row].height = 30
                

                # Only add background color for self-assessment cells
                if i == j:  # Use indices instead of model names to handle duplicates
                    cell.fill = SELF_ASSESSMENT_FILL
            
            # Calculate and add median grade
            median_grade, _, median_value, raw_average = calculate_median_grade(
                grades_for_model, 
                results, graded_model, grader_models_for_median
            )
            
            # Calculate raw average directly to ensure maximum precision
            exact_raw_average = 0.0
            if raw_numeric_grades:
                exact_raw_average = sum(raw_numeric_grades) / len(raw_numeric_grades)
                
            # Add median grade with its numeric equivalent
            cell = ws.cell(row=row, column=median_col)
            
            # Format median for display
            median_display = f"{raw_average:.2f}" if raw_average > 0 else "N/A"
            
            # Add composite grade notation if applicable
            if raw_average > 0:
                composite_grade = numeric_to_composite_grade(raw_average)
                median_display = composite_grade
            
            apply_cell_style( 
                cell,
                value=median_display,
                alignment=CENTER_WRAP_ALIGNMENT,
                border=THIN_BORDER
            )
            
            # Add the exact raw average to the raw average column with 4 decimal precision
            exact_raw_display = f"{exact_raw_average:.2f}" if exact_raw_average > 0 else "N/A"
            
            # Apply composite grade to raw average if applicable
            if exact_raw_average > 0:
                composite_grade = numeric_to_composite_grade(exact_raw_average)
                if "/" in composite_grade:  # If it's a composite grade
                    exact_raw_display = f"{composite_grade} ({exact_raw_average:.2f})"
                else:
                    exact_raw_display = f"{composite_grade}"
            
            apply_cell_style(
                ws.cell(row=row, column=raw_avg_col),
                value=exact_raw_display,
                alignment=CENTER_WRAP_ALIGNMENT,
                border=THIN_BORDER,
                font=Font(name="Helvetica Neue", bold=True, size=8),  # Make the raw average prominent with 8pt
                fill=PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
            )
            
        # Add Grading Bias row
        bias_row = len(models) + 3
        apply_cell_style(
            ws.cell(row=bias_row, column=1),
            value="Grading Bias (Lenient/Strict Tendencies)",
            alignment=WRAP_ALIGNMENT,
            border=THIN_BORDER,
            fill=BIAS_ROW_FILL,
            font=HELVETICA_BOLD
        )
        
        # Fill in grading bias data with improved clarity
        if "bias_analysis" in results and "grader_bias" in results["bias_analysis"]:
            for j, grader_model in enumerate(models):
                col = j + 2
                
                bias_text = ""  
                if (grader_model in results["bias_analysis"]["grader_bias"] and 
                    "letter_bias" in results["bias_analysis"]["grader_bias"][grader_model]):
                    bias_text = results["bias_analysis"]["grader_bias"][grader_model]["letter_bias"]
                
                apply_cell_style(
                    ws.cell(row=bias_row, column=col), 
                    value=bias_text,
                    alignment=CENTER_WRAP_ALIGNMENT,
                    border=THIN_BORDER,
                    fill=BIAS_ROW_FILL
                )
            
            # Add empty cell for median column in bias row
            apply_cell_style(
                ws.cell(row=bias_row, column=median_col),
                value="N/A",
                border=THIN_BORDER,
                fill=BIAS_ROW_FILL
            )
            
            # Add empty cell for raw average column in bias row
            apply_cell_style(
                ws.cell(row=bias_row, column=raw_avg_col),
                value="N/A",
                border=THIN_BORDER,
                fill=BIAS_ROW_FILL
            )
        
        # Add Column Median row
        col_median_row = len(models) + 4
        apply_cell_style(
            ws.cell(row=col_median_row, column=1),
            value="Median Grade Given by Model",
            alignment=WRAP_ALIGNMENT,
            border=THIN_BORDER,
            fill=HEADER_ROW_FILL,
            font=HELVETICA_BOLD
        )
        
        # Fill in column median data
        for j, grader_model in enumerate(models):
            col = j + 2
            # Get display name for this grader 
            display_grader_model = display_models[j]
            
            # Calculate median for this column
            # Collect all grades given by this grader model
            grades_given = []
            for graded_model in models:
                try:
                    grade_info = results["grades"][graded_model][grader_model]
                    if isinstance(grade_info, dict) and "grade" in grade_info and grade_info["grade"] != "N/A":
                        grades_given.append(grade_info["grade"])
                except KeyError:
                    pass
            
            # Calculate median
            median_grade = "N/A"
            median_value = "N/A"
            display_value = "N/A"
            median_score = 0.0
            descriptive_text = "N/A"
            
            if grades_given:
                median_grade, median_score, median_value, _ = calculate_median_grade(grades_given)
                
                # Apply composite grade notation if applicable
                if median_score > 0:
                    composite_grade = numeric_to_composite_grade(median_score)
                    display_value = composite_grade
                
                # Get descriptive text based on median grade
                if median_grade == "A+":
                    descriptive_text = "Excellent"
                elif median_grade in ["A", "A-"]:
                    descriptive_text = "Very Good"
                elif median_grade in ["B+", "B"]:
                    descriptive_text = "Good"
                elif median_grade in ["B-", "C+"]:
                    descriptive_text = "Satisfactory"
                elif median_grade in ["C", "C-"]:
                    descriptive_text = "Average"
                elif median_grade.startswith("D"):
                    descriptive_text = "Below Average"
                elif median_grade == "F":
                    descriptive_text = "Poor"
            else:
                display_value = "N/A\n(No Data)"
            
            apply_cell_style(
                ws.cell(row=col_median_row, column=col),
                value=display_value,
                alignment=CENTER_WRAP_ALIGNMENT,
                border=THIN_BORDER,
                fill=HEADER_ROW_FILL,
                font=HELVETICA_BOLD
            )
        
        # Add empty cells for median and raw average columns
        apply_cell_style(
            ws.cell(row=col_median_row, column=median_col),
            value="",
            border=THIN_BORDER,
            fill=HEADER_ROW_FILL
        )
        
        apply_cell_style(
            ws.cell(row=col_median_row, column=raw_avg_col),
            value="",
            border=THIN_BORDER,
            fill=HEADER_ROW_FILL
        )
        
        # Add legend explaining the table at the bottom
        legend_row = len(models) + 6
        apply_cell_style(
            ws.cell(row=legend_row, column=1, value="Table Legend:"),
            alignment=LEFT_ALIGNMENT,
            font=Font(name="Helvetica Neue", bold=True, size=8)
        )
        
        # Expand legend with more detailed explanation
        apply_cell_style(
            ws.cell(row=legend_row+1, column=1),
            value="- Rows: Each row represents a model being evaluated (essay author)",
            alignment=LEFT_ALIGNMENT,
            font=HELVETICA_FONT
        )
        
        apply_cell_style(
            ws.cell(row=legend_row+2, column=1),
            value="- Columns: Each column shows the score given by a particular grading model",
            alignment=LEFT_ALIGNMENT,
            font=HELVETICA_FONT
        )
        
        apply_cell_style(
            ws.cell(row=legend_row+3, column=1),
            value="- Self-assessments (where a model grades its own essay) are highlighted in lavender",
            alignment=LEFT_ALIGNMENT,
            font=HELVETICA_FONT
        )
        
        apply_cell_style(
            ws.cell(row=legend_row+4, column=1),
            value="- Overall Score column shows the numerical average of all grades on a 0-4.25 scale",
            alignment=LEFT_ALIGNMENT,
            font=HELVETICA_FONT
        )
                
        # Add model name appendix
        appendix_row = legend_row + 6
        if model_name_mapping:
            apply_cell_style(
                ws.cell(row=appendix_row, column=1),
                value="Model Name Reference:",
                alignment=LEFT_ALIGNMENT,
                font=Font(name="Helvetica Neue", bold=True, size=8)
            )
            
            # Add mapping between abbreviated and full names
            for i, (abbrev, full_name) in enumerate(model_name_mapping.items()):
                if abbrev != full_name:  # Only add if there was an actual abbreviation
                    apply_cell_style(
                        ws.cell(row=appendix_row + i + 1, column=1),
                        value=f"{abbrev} = {full_name}",
                        alignment=LEFT_ALIGNMENT,
                        font=HELVETICA_FONT
                    )
        
        # Extract and save raw grade matrix for future use
        matrix = extract_raw_grade_matrix(results)
        matrix_file = os.path.join(run_dir, "raw_grade_matrix.json")
        with open(matrix_file, 'w') as f:
            json.dump(matrix, f, indent=2)
        print(f"Saved raw grade matrix to: {matrix_file}")
        
        # Save the Excel file
        excel_file = os.path.join(run_dir, "cross_grading_table.xlsx")
        ws.title = "Cross Grading (Letter Only)"
        wb.save(excel_file)
        print(f"Generated enhanced Excel cross-grading table: {excel_file}")
        
    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        if isinstance(e, (KeyError, IndexError, ValueError, TypeError)):
            # More specific error handling for common issues
            print(f"Details: {type(e).__name__} - Check the structure of your results dictionary")
        # Optionally re-raise the exception if needed
        # raise