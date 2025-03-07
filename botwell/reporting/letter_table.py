"""
Letter grade table generation for Boswell Test results.

This module generates a simplified Excel spreadsheet with letter grades
without color coding or extra formatting, matching the format in the reference.
"""

import os
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from botwell.reporting.excel import abbreviate_model_name 
from botwell.core.grading import grade_to_numeric, numeric_to_composite_grade
from statistics import median

# Define minimal styling constants
THIN_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'),
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
BOLD_FONT = Font(name="Helvetica Neue", bold=True, size=8)
REGULAR_FONT = Font(name="Helvetica Neue", size=8)
CENTER_ALIGNMENT = Alignment(horizontal='center')

def calculate_median_letter_grade(letter_grades: list) -> str:
    """
    Calculate median grade from a list of letter grades.
    
    Args:
        letter_grades: List of letter grade strings (e.g., ["A", "B+", "A-"])
        
    Returns:
        Median letter grade
    """
    if not letter_grades:
        return "N/A"
    
    # Special case for identical grades
    if all(grade == letter_grades[0] for grade in letter_grades):
        return letter_grades[0]
        
    # Convert to numeric scores
    numeric_grades = [grade_to_numeric(grade) for grade in letter_grades]
    
    # Calculate median
    numeric_grades.sort()
    length = len(numeric_grades)
    if length % 2 == 0:  # Even number, average the two middle values
        median_score = (numeric_grades[length//2 - 1] + numeric_grades[length//2]) / 2
    else:  # Odd number, take the middle value
        median_score = numeric_grades[length//2]
    
    # Find corresponding letter grade
    # Use the new composite grade function to get more precise grade notation
    return numeric_to_composite_grade(median_score)

def generate_letter_grade_table(results: Dict[str, Any], output_path: str) -> str:
    """
    Generate a simplified Excel table with letter grades and BQ Numeric Average.
    
    Args:
        results: The full results dictionary with grades
        output_path: Path where to save the Excel file
        
    Returns:
        Path to the generated Excel file
    """
    try:
        # Extract models list
        models = list(results["essays"].keys())
        if not models:
            raise ValueError("No models found in results")
            
        # Abbreviate model names for display
        display_models = [abbreviate_model_name(model) for model in models]
        
        # Create workbook and get active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Letter Grade Data"
        
        # Set up header row
        ws.cell(row=1, column=1, value="Model").font = BOLD_FONT
        ws.cell(row=1, column=1).border = THIN_BORDER
        
        # Add model columns
        for j, model in enumerate(display_models):
            cell = ws.cell(row=1, column=j+2, value=model)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            # Set narrow column width for compact display
            ws.column_dimensions[get_column_letter(j+2)].width = 10
            
        # Add median and BQ numeric average columns
        median_col = len(models) + 2
        bq_avg_col = len(models) + 3
        
        ws.cell(row=1, column=median_col, value="Median Grade").font = BOLD_FONT
        ws.cell(row=1, column=median_col).border = THIN_BORDER
        
        ws.cell(row=1, column=bq_avg_col, value="BQ Numeric Average").font = BOLD_FONT
        ws.cell(row=1, column=bq_avg_col).border = THIN_BORDER
        
        # Fill in letter grade data rows
        for i, graded_model in enumerate(models):
            row = i + 2  # Start on row 2
            
            # Add model name in first column
            model_cell = ws.cell(row=row, column=1, value=display_models[i])
            model_cell.font = BOLD_FONT
            model_cell.border = THIN_BORDER
            
            # Collect all grades for this model
            letter_grades = []
            numeric_grades = []
            
            # Add each grade as letter grade
            for j, grader_model in enumerate(models):
                col = j + 2
                
                # Extract grade data
                try:
                    grade_info = results["grades"][graded_model][grader_model]
                    if isinstance(grade_info, dict) and "grade" in grade_info:
                        letter_grade = grade_info["grade"]
                        if letter_grade != "N/A":
                            # Add letter grade to cell
                            ws.cell(row=row, column=col, value=letter_grade)
                            letter_grades.append(letter_grade)
                            
                            # Get numeric grade (either from stored value or convert from letter)
                            if "numeric_grade" in grade_info:
                                numeric_grade = grade_info["numeric_grade"]
                            else:
                                numeric_grade = grade_to_numeric(letter_grade)
                            
                            numeric_grades.append(numeric_grade)
                except (KeyError, TypeError):
                    # Leave cell empty if no grade data
                    pass
                
                # Add border to all cells
                ws.cell(row=row, column=col).border = THIN_BORDER
            
            # Calculate and add median letter grade
            if letter_grades:
                median_letter = calculate_median_letter_grade(letter_grades)
                ws.cell(row=row, column=median_col, value=median_letter)
            
            # Calculate and add BQ numeric average
            if numeric_grades:
                bq_avg = sum(numeric_grades) / len(numeric_grades)
                # Add the BQ numeric average with appropriate precision
                numeric_value = round(bq_avg, 2)
                ws.cell(row=row, column=bq_avg_col, value=numeric_value)
                
                # Optionally, add the equivalent composite letter grade in a note
                # This helps users understand the exact meaning of the numeric value
                composite_grade = numeric_to_composite_grade(bq_avg)
                if "/" in composite_grade:  # It's a composite grade
                    cell = ws.cell(row=row, column=bq_avg_col)
                    cell.comment = openpyxl.comments.Comment(
                        f"Equivalent to composite grade: {composite_grade}", "Boswell System"
                    )
            
            # Add borders
            ws.cell(row=row, column=median_col).border = THIN_BORDER
            ws.cell(row=row, column=bq_avg_col).border = THIN_BORDER
        
        # Add column median row at the bottom
        col_median_row = len(models) + 2
        median_label = ws.cell(row=col_median_row, column=1, value="Median Grade Given")
        median_label.font = BOLD_FONT
        median_label.border = THIN_BORDER
        
        # Calculate and add column medians (letter grades given by each model)
        for j, grader_model in enumerate(models):
            col = j + 2
            
            # Collect all letter grades given by this model
            col_grades = []
            for graded_model in models:
                try:
                    if graded_model in results["grades"] and grader_model in results["grades"][graded_model]:
                        grade_info = results["grades"][graded_model][grader_model]
                        if isinstance(grade_info, dict) and "grade" in grade_info:
                            letter_grade = grade_info["grade"]
                            if letter_grade != "N/A":
                                col_grades.append(letter_grade)
                except KeyError:
                    pass
            
            # Calculate and add median for this column
            if col_grades:
                median_letter = calculate_median_letter_grade(col_grades)
                ws.cell(row=col_median_row, column=col, value=median_letter)
            
            # Apply border to cells
            ws.cell(row=col_median_row, column=col).border = THIN_BORDER
        
        # Add borders to median and BQ avg cells in bottom row
        ws.cell(row=col_median_row, column=median_col).border = THIN_BORDER
        ws.cell(row=col_median_row, column=bq_avg_col).border = THIN_BORDER
        
        # Save the Excel file
        excel_file = os.path.join(output_path, "letter_grade_table.xlsx")
        wb.save(excel_file)
        
        return excel_file
        
    except Exception as e:
        print(f"Error generating letter grade table: {str(e)}")
        if isinstance(e, (KeyError, IndexError, ValueError, TypeError)):
            print(f"Details: {type(e).__name__} - Check the data structure")
        return None